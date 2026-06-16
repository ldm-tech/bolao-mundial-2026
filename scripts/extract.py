"""
Extrai o fixture oficial + os palpites de cada jogador das planilhas .xlsx
para um JSON limpo (data/seed.json). Roda uma vez; o app Node consome o JSON.

Casa tudo pelo NUMERO DO JOGO (estavel entre planilhas), nunca pelo nome do time.
"""
import json
import re
from pathlib import Path
import openpyxl

PLANILHAS = Path(__file__).resolve().parent.parent / "planilhas"
OUT = Path(__file__).resolve().parent.parent / "data" / "seed.json"

# arquivo -> nome de exibicao (com acento)
JOGADORES = {
    "Danilo": "Danilo", "Fabio": "Fabio", "Felipe": "Felipe", "Julia": "Julia",
    "Manu": "Manu", "Rebeca": "Rebeca", "Renata": "Renata", "Theo": "Theo", "Tim": "Tim",
}

import re as _re

# Limpa o nome do arquivo para virar nome de exibicao quando a planilha
# nao teve o campo NOME preenchido ("Seu Nome", o default do template).
def nome_do_arquivo(stem):
    s = stem
    s = _re.sub(r'(?i)bol[aã]o\s*pedreira', ' ', s)
    s = _re.sub(r'(?i)\bbol[aã]o\b', ' ', s)
    s = s.replace('_', ' ').replace('-', ' ').replace('.', ' ')
    s = _re.sub(r'\s+', ' ', s).strip()
    s = s.strip('()[] ').strip()
    return s or stem

# Le NOME/EMAIL/WHATSAPP da aba Regras (G21..G23 = col 7, linhas 21-23).
def le_contato(wb):
    ws = wb['Regras']
    nome = ws.cell(row=21, column=7).value
    email = ws.cell(row=22, column=7).value
    whats = ws.cell(row=23, column=7).value
    norm = lambda v: (str(v).strip() if v is not None and str(v).strip() != '' else None)
    return norm(nome), norm(email), norm(whats)

NOME_PLACEHOLDER = 'Seu Nome'

# Padroniza caps: primeira maiuscula por palavra, resto minusculo. Conectivos
# em PT (de, da, e...) ficam minusculos quando nao sao a 1a palavra.
_CONECTIVOS = {'de', 'da', 'do', 'das', 'dos', 'e', 'di', 'du', 'del', 'la',
               'o', 'a', 'os', 'as', 'van', 'von', 'y'}

def normaliza_caps(nome):
    if not nome:
        return nome
    palavras = nome.split()
    out = []
    for i, p in enumerate(palavras):
        baixo = p.lower()
        if i > 0 and baixo in _CONECTIVOS:
            out.append(baixo)
        else:
            out.append(baixo[:1].upper() + baixo[1:])
    return ' '.join(out)


def fase_do_jogo(n):
    if n <= 72: return "grupos"
    if n <= 88: return "1/16"
    if n <= 96: return "oitavas"
    if n <= 100: return "quartas"
    if n <= 102: return "semis"
    if n == 103: return "terceiro"
    return "final"


def parse_cidade_pais(texto):
    if not texto:
        return (None, None)
    m = re.match(r"^(.*?)\s*\(([^)]+)\)\s*$", str(texto).strip())
    if m:
        return (m.group(1).strip(), m.group(2).strip())
    return (str(texto).strip(), None)


def to_int(v):
    if v is None or str(v).strip() == "":
        return None
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return None


def extrai_jogos_pontuacao(ws):
    """Le a aba PONTUACAO: para cada jogo numerado retorna metadados + placar palpitado."""
    jogos = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        numero = to_int(row[1]) if len(row) > 1 else None  # col 2
        if numero is None or numero < 1 or numero > 104:
            continue
        data = row[2] if len(row) > 2 else None             # col 3
        hora = row[3] if len(row) > 3 else None             # col 4
        cidade_pais = row[4] if len(row) > 4 else None       # col 5
        time_casa = row[6] if len(row) > 6 else None         # col 7
        gols_casa = to_int(row[7]) if len(row) > 7 else None # col 8
        gols_fora = to_int(row[9]) if len(row) > 9 else None # col 10
        time_fora = row[10] if len(row) > 10 else None       # col 11
        cidade, pais = parse_cidade_pais(cidade_pais)
        jogos[numero] = {
            "numero": numero,
            "fase": fase_do_jogo(numero),
            "data": str(data)[:10] if data else None,
            "hora": str(hora)[:5] if hora else None,
            "cidade": cidade,
            "pais": pais,
            "time_casa": str(time_casa).strip() if time_casa else None,
            "time_fora": str(time_fora).strip() if time_fora else None,
            "gols_casa": gols_casa,
            "gols_fora": gols_fora,
        }
    return jogos


# A aba PONTUACAO calcula os gols de grupos por formula (=PALPITES!Xn). Quando o
# palpite foi deixado EM BRANCO, a formula vira 0 e (com data_only) era importado
# como 0 — dando pontos indevidos. Aqui seguimos a formula ate a celula crua do
# PALPITES p/ saber se o palpite estava de fato vazio. Precisa do wb com formulas
# (data_only=False).
_GRUPO_REF = re.compile(r"PALPITES!\$?([A-Z]+)\$?(\d+)", re.I)


def _palpite_vazio(cell, palp):
    f = cell.value
    if isinstance(f, str) and f.startswith("="):
        m = _GRUPO_REF.search(f)
        if m:
            return palp[m.group(1).upper() + m.group(2)].value is None
    return False  # valor literal digitado direto: nao e derivado de branco


def grupos_brancos(wb_formulas):
    """{numero: (casa_vazio, fora_vazio)} p/ os jogos de grupos (1..72)."""
    pontu = wb_formulas["PONTUAÇÃO"]
    palp = wb_formulas["PALPITES"]
    brancos = {}
    for row in pontu.iter_rows(min_row=1, max_row=pontu.max_row):
        numero = to_int(row[1].value)
        if numero is None or not (1 <= numero <= 72):
            continue
        brancos[numero] = (_palpite_vazio(row[7], palp), _palpite_vazio(row[9], palp))  # col 8, col 10
    return brancos


def extrai_especiais(wb):
    """Artilheiro (PALPITES L106) e campeao (PALPITES R98). Finalistas vem do jogo 104."""
    ws = wb["PALPITES"]
    artilheiro = ws.cell(row=106, column=12).value
    campeao = ws.cell(row=98, column=18).value
    return {
        "artilheiro": str(artilheiro).strip() if artilheiro else None,
        "campeao": str(campeao).strip() if campeao else None,
    }


# Normaliza nome de selecao para casar confrontos entre as abas PALPITES e PONTUACAO.
_SINONIMOS = {
    "qatar": "catar", "rep tcheca": "republica tcheca", "tcheca": "republica tcheca",
    "bosnia hezerg": "bosnia", "bosnia herzegovina": "bosnia", "curacau": "curacao",
    "rd congo": "congo", "eua": "estados unidos",
}


def norm_time(s):
    if not s:
        return None
    import unicodedata
    s = unicodedata.normalize("NFD", str(s))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn").lower()
    s = s.replace(".", "").replace("-", " ")
    s = " ".join(s.split())
    return _SINONIMOS.get(s, s) or None


# Rodadas do mata-mata na aba PALPITES. Cada bloco de jogo comeca na coluna do
# rotulo numerico; times = (numcol+1, numcol+fora_off); goalrow = numrow+3; penrow = numrow+5.
# A NUMERACAO do PALPITES difere da PONTUACAO, entao casamos por CONFRONTO, nao por numero.
MATA_ROUNDS = [
    {"numrow": 48, "fora_off": 2},  # 1/16
    {"numrow": 58, "fora_off": 2},  # oitavas
    {"numrow": 68, "fora_off": 3},  # quartas
    {"numrow": 78, "fora_off": 3},  # semis
    {"numrow": 86, "fora_off": 3},  # terceiro lugar
    {"numrow": 96, "fora_off": 3},  # final
]


def extrai_mata_palpites(wb):
    """Le os confrontos do mata-mata no PALPITES (fonte autoritativa).
    Retorna {frozenset(times_norm): {home, gols(h,a), pen(h,a)}}."""
    ws = wb["PALPITES"]
    confrontos = {}
    for r in MATA_ROUNDS:
        numrow, teamrow, goalrow, penrow = r["numrow"], r["numrow"] + 2, r["numrow"] + 3, r["numrow"] + 5
        for col in range(1, ws.max_column + 1):
            numero = to_int(ws.cell(row=numrow, column=col).value)
            if numero is None or not (73 <= numero <= 104):
                continue  # so colunas que rotulam um jogo de mata-mata
            home = ws.cell(row=teamrow, column=col + 1).value
            away = ws.cell(row=teamrow, column=col + r["fora_off"]).value
            if not home or not away:
                continue
            nh, na = norm_time(home), norm_time(away)
            confrontos[frozenset((nh, na))] = {
                "home": nh,
                "gols": (to_int(ws.cell(row=goalrow, column=col + 1).value),
                         to_int(ws.cell(row=goalrow, column=col + r["fora_off"]).value)),
                "pen": (to_int(ws.cell(row=penrow, column=col + 1).value),
                        to_int(ws.cell(row=penrow, column=col + r["fora_off"]).value)),
            }
    return confrontos


def casa_confronto(confrontos, time_casa, time_fora):
    """Busca o confronto do PALPITES e orienta (gols/pen) conforme o mando dado."""
    nh, na = norm_time(time_casa), norm_time(time_fora)
    info = confrontos.get(frozenset((nh, na)))
    if not info:
        return None
    gc, gf = info["gols"]
    pc, pf = info["pen"]
    if info["home"] != nh:  # mando invertido entre as abas -> troca os pares
        gc, gf = gf, gc
        pc, pf = pf, pc
    return {"gols_casa": gc, "gols_fora": gf, "pen_casa": pc, "pen_fora": pf}


# Classificacao prevista de cada grupo (PALPITES, linhas 42-45: 1o a 4o lugar).
# Os codigos "1A".."4L" trazem posicao + grupo; o time fica na coluna ao lado.
def extrai_grupos(wb):
    ws = wb["PALPITES"]
    out = {}  # letra -> [1o, 2o, 3o, 4o]
    for row in (42, 43, 44, 45):
        for c in range(1, ws.max_column + 1):
            code = ws.cell(row=row, column=c).value
            if not code:
                continue
            m = re.match(r"^([1-4])([A-L])$", str(code).strip())
            if not m:
                continue
            pos, letra = int(m.group(1)), m.group(2)
            time = ws.cell(row=row, column=c + 1).value
            out.setdefault(letra, [None, None, None, None])[pos - 1] = (
                str(time).strip() if time else None
            )
    return out


def main():
    arquivos = sorted(PLANILHAS.glob("*.xlsx"))
    fixture = None
    jogadores = []
    contatos = []
    palpites = {}      # nome -> {numero: {gols_casa, gols_fora, time_casa, time_fora}}
    especiais = {}
    grupos = {}        # nome -> {letra: [1o, 2o, 3o, 4o]}

    for arq in arquivos:
        chave = arq.stem  # chave estavel (nome do arquivo), nao muda no admin
        wb = openpyxl.load_workbook(arq, data_only=True)
        wb_formulas = openpyxl.load_workbook(arq, data_only=False)  # p/ ler as formulas dos grupos
        brancos = grupos_brancos(wb_formulas)  # palpites de grupos deixados em branco
        nome_campo, email, whats = le_contato(wb)
        nome = nome_campo if (nome_campo and nome_campo != NOME_PLACEHOLDER) else nome_do_arquivo(arq.stem)
        nome = normaliza_caps(nome)  # padroniza MAIUSCULAS/minusculas
        print(f"Lendo {arq.name} -> {nome}")
        ws = wb["PONTUAÇÃO"]
        jogos = extrai_jogos_pontuacao(ws)

        # fixture oficial (metadados + times dos grupos) vem do primeiro arquivo
        if fixture is None:
            fixture = {}
            for n, j in jogos.items():
                fixture[n] = {
                    "numero": n, "fase": j["fase"], "data": j["data"], "hora": j["hora"],
                    "cidade": j["cidade"], "pais": j["pais"],
                    # so grupos tem time oficial; mata-mata depende do resultado real
                    "time_casa": j["time_casa"] if j["fase"] == "grupos" else None,
                    "time_fora": j["time_fora"] if j["fase"] == "grupos" else None,
                }

        jogadores.append(nome)
        contatos.append({"chave": chave, "nome": nome, "email": email, "whatsapp": whats})
        confrontos = extrai_mata_palpites(wb)
        pdict = {}
        for n, j in jogos.items():
            gc, gf, pen_c, pen_f = j["gols_casa"], j["gols_fora"], None, None
            if n >= 73:  # mata-mata: pega penaltis (e corrige o 3o lugar) pelo PALPITES
                m = casa_confronto(confrontos, j["time_casa"], j["time_fora"])
                if m:
                    pen_c, pen_f = m["pen_casa"], m["pen_fora"]
                    # jogo 103 (3o lugar) tem a PONTUACAO obsoleta no template: usa o PALPITES
                    if n == 103 and m["gols_casa"] is not None and m["gols_fora"] is not None:
                        gc, gf = m["gols_casa"], m["gols_fora"]
            else:  # grupos: respeita o palpite EM BRANCO (a formula da PONTUACAO vira 0)
                bc, bf = brancos.get(n, (False, False))
                if bc:
                    gc = None
                if bf:
                    gf = None
            pdict[str(n)] = {
                "gols_casa": gc, "gols_fora": gf,
                "time_casa": j["time_casa"], "time_fora": j["time_fora"],
                "pen_casa": pen_c, "pen_fora": pen_f,
            }
        palpites[nome] = pdict
        esp = extrai_especiais(wb)
        # finalistas = times do jogo 104 (final) no palpite do jogador
        final = jogos.get(104, {})
        esp["finalista_1"] = final.get("time_casa")
        esp["finalista_2"] = final.get("time_fora")
        especiais[nome] = esp
        grupos[nome] = extrai_grupos(wb)

    dups = [n for n in set(jogadores) if jogadores.count(n) > 1]
    if dups:
        raise SystemExit(f"ERRO: nomes de exibicao repetidos {dups}. "
                         f"Renomeie o arquivo ou ajuste manualmente antes de prosseguir.")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    seed = {
        "jogadores": jogadores,
        "fixture": [fixture[n] for n in sorted(fixture)],
        "palpites": palpites,
        "especiais": especiais,
        "grupos": grupos,
    }
    OUT.write_text(json.dumps(seed, ensure_ascii=False, indent=2), encoding="utf-8")
    CONTATOS = OUT.parent / "contatos.json"
    CONTATOS.write_text(json.dumps(contatos, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\nOK: {len(jogadores)} jogadores, {len(fixture)} jogos -> {OUT}")
    print(f"Contatos (PII, NAO committar): {len(contatos)} -> {CONTATOS}")


if __name__ == "__main__":
    main()
