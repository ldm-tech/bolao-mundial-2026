"""Varredura: acha palpites de GRUPOS deixados EM BRANCO na planilha que a
PONTUACAO (com formula) importou como 0 (porque =PALPITES!Xn de celula vazia
vira 0). Segue a formula ate a celula crua do PALPITES e checa se esta vazia.
Nao altera nada — so relata.
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

PASTA = Path("planilhas")
SEED = Path("data/seed.json")
REF = re.compile(r"PALPITES!\$?([A-Z]+)\$?(\d+)", re.I)

# nomes oficiais dos jogos (so p/ leitura do relatorio)
fixture = {}
if SEED.exists():
    seed = json.loads(SEED.read_text(encoding="utf-8"))
    fx = seed.get("fixture") or {}
    itens = fx.values() if isinstance(fx, dict) else fx
    for j in itens:
        try:
            fixture[int(j.get("numero"))] = (j.get("time_casa"), j.get("time_fora"))
        except (TypeError, ValueError):
            pass


def nome_de(wb, stem):
    try:
        v = wb["Regras"].cell(row=21, column=7).value
        if v and str(v).strip() and str(v).strip().lower() != "seu nome":
            return str(v).strip()
    except Exception:
        pass
    return stem


casos = []
arqs = sorted(PASTA.glob("*.xlsx"))
for i, arq in enumerate(arqs, 1):
    print(f"[{i}/{len(arqs)}] {arq.name}", file=sys.stderr)
    try:
        wb = openpyxl.load_workbook(arq, data_only=False)
    except Exception as e:
        print("ERRO load", arq.name, e, file=sys.stderr)
        continue
    try:
        pontu = wb["PONTUAÇÃO"]
        palp = wb["PALPITES"]
        nome = nome_de(wb, arq.stem)
        for row in pontu.iter_rows(min_row=1, max_row=pontu.max_row):
            try:
                numero = int(float(row[1].value))
            except (TypeError, ValueError):
                continue
            if not (1 <= numero <= 72):  # so fase de grupos
                continue

            def raw(cell_obj):
                f = cell_obj.value
                if isinstance(f, str) and f.startswith("="):
                    m = REF.search(f)
                    if m:
                        return palp[m.group(1).upper() + m.group(2)].value
                    return "??"  # formula nao reconhecida (nao aponta p/ PALPITES)
                return f  # valor literal digitado direto

            rc = raw(row[7])   # col 8 = gols_casa
            rf = raw(row[9])   # col 10 = gols_fora
            cblk = rc is None
            fblk = rf is None
            if cblk or fblk:
                casos.append((numero, nome, arq.stem, cblk, fblk))
    finally:
        wb.close()

print("\n=== PALPITES DE GRUPOS COM CELULA EM BRANCO (importados como 0) ===")
total = len(casos)
vazio_total = [c for c in casos if c[3] and c[4]]
print(f"total de celulas afetadas: {total}")
print(f"  VAZIO TOTAL (importou 0x0 — pode ter dado pontos errados): {len(vazio_total)}")
print(f"  parcial (so uma celula em branco): {total - len(vazio_total)}")
print("\n--- detalhe (ordenado por nome) ---")
for numero, nome, stem, cblk, fblk in sorted(casos, key=lambda x: (x[1], x[0])):
    tc, tf = fixture.get(numero, ("?", "?"))
    tipo = "VAZIO TOTAL (0x0)" if (cblk and fblk) else ("falta casa" if cblk else "falta fora")
    print(f"  jogo {numero:>2} {tc} x {tf}  |  {nome}  [{stem}]  -> {tipo}")
